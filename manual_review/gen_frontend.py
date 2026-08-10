# -*- coding: utf-8 -*-
"""人工核查系统前端生成器（卡片式）
从《配置核查表_v2.0.0.xlsx》抽取全部核查项，生成自包含 HTML 人工核查界面。
产物：核查系统.html —— 双击即可在浏览器打开，离线录入、保存进度、导出 Excel 报告。
"""
import json
import os
import re
import openpyxl

BASE = r"c:\Users\ryan.xiong\Desktop\配置核查"
XLSX = BASE + r"\配置核查表_v2.0.0.xlsx"
OUT = BASE + r"\manual_review\核查系统.html"

CHAP_NO = {
    "系统安全": 1, "用户安全": 2, "数据安全": 3, "应用安全": 4,
    "网络安全": 5, "物理安全": 6, "组织机构": 7, "规章制度": 8,
    "管理实施": 9, "协议安全审计": 10,
}
OBJ_COLS = ["Win7", "WinXP", "中标麒麟", "银河麒麟", "Nginx", "Tomcat",
            "MySQL", "SQL Server", "达梦", "Redis", "其他"]


def norm(text):
    s = str(text).replace("\n", " ").replace("\r", "")
    s = re.sub(r"^\s*\d+\s*[、.．]\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_items():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]
    chapter = ""
    items = []
    counters = {}
    for r in range(4, 138):
        a = ws.cell(r, 1).value
        if a and str(a).strip():
            chapter = str(a).strip()
        item = ws.cell(r, 2).value
        if not item or not str(item).strip():
            continue
        cno = CHAP_NO.get(chapter, 0)
        counters[chapter] = counters.get(chapter, 0) + 1
        idx = counters[chapter]
        eid = f"{cno}.{idx}" if cno else f"?.{idx}"
        plats = []
        for j, name in enumerate(OBJ_COLS):
            v = ws.cell(r, 3 + j).value
            if v and "√" in str(v):
                plats.append(name)
        items.append({
            "id": eid,
            "chapter": chapter,
            "cno": cno,
            "category": chapter,
            "title": norm(item),
            "guide": f"《配置核查作业指导书》第{cno}章 {chapter}" if cno else "",
            "plat": "、".join(plats),
        })
    return items


HTML = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>配置核查 - 人工核查系统</title>
<style>
*{box-sizing:border-box;}
body{font-family:"Microsoft YaHei","微软雅黑",Arial,sans-serif;margin:0;background:#f0f2f5;color:#222;font-size:13px;}
header{background:linear-gradient(135deg,#1a3c6e,#2c5aa0);color:#fff;padding:14px 22px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 6px rgba(0,0,0,.15);}
header h1{margin:0;font-size:18px;font-weight:600;}
header .actions{display:flex;gap:8px;}
header .actions button{background:#fff;color:#1a3c6e;border:none;padding:8px 14px;border-radius:4px;font-size:13px;font-weight:600;cursor:pointer;}
header .actions button:hover{background:#e8eef7;}
header .actions button.danger{color:#c62828;}
.bar{background:#fff;padding:10px 22px;border-bottom:1px solid #e0e0e0;display:flex;flex-wrap:wrap;gap:10px;align-items:center;}
.bar label{font-size:12px;color:#555;}
.bar input{padding:5px 8px;border:1px solid #ccc;border-radius:3px;font-size:12px;width:140px;}
.stats{background:#fff;padding:10px 22px;border-bottom:1px solid #e0e0e0;display:flex;gap:10px;align-items:center;flex-wrap:wrap;}
.stats .chip{padding:6px 12px;border-radius:14px;font-size:12px;font-weight:600;color:#fff;}
.stats .s-pass{background:#2e7d32;} .stats .s-fail{background:#c62828;}
.stats .s-manual{background:#ef6c00;} .stats .s-na{background:#757575;} .stats .s-empty{background:#b0bec5;color:#333;}
.stats .progress{margin-left:auto;font-size:12px;color:#555;}
.stats .progress b{color:#1a3c6e;}
.filters{background:#fff;padding:8px 22px;border-bottom:1px solid #e0e0e0;display:flex;gap:6px;flex-wrap:wrap;align-items:center;position:sticky;top:0;z-index:10;}
.filters .tab{padding:5px 11px;border:1px solid #ccc;border-radius:14px;cursor:pointer;font-size:12px;background:#f7f9fb;}
.filters .tab.active{background:#1a3c6e;color:#fff;border-color:#1a3c6e;}
.filters .tab .n{color:#888;font-size:11px;margin-left:3px;}
.filters .tab.active .n{color:#cfe;}
.filters input[type=text]{margin-left:auto;padding:5px 10px;border:1px solid #ccc;border-radius:3px;font-size:12px;width:200px;}
textarea{width:100%;border:1px solid #e0e0e0;border-radius:3px;padding:5px 7px;font-size:12px;font-family:inherit;resize:vertical;min-height:36px;background:#fafbfc;}
textarea:focus{border-color:#2c5aa0;background:#fff;outline:none;}
.banner{background:#fff3cd;color:#856404;padding:6px 22px;font-size:12px;border-bottom:1px solid #ffe69e;display:none;}
.banner.show{display:block;}
.batchbar{background:#eef2f7;padding:8px 22px;border-bottom:1px solid #d0d8e4;display:flex;gap:8px;align-items:center;font-size:12px;flex-wrap:wrap;}
.batchbar .lbl{color:#1a3c6e;font-weight:600;}
.batchbar button{border:1px solid #bbb;background:#fff;padding:4px 9px;border-radius:3px;cursor:pointer;font-size:11px;}
.batchbar button:hover{background:#e8eef7;}
.batchbar .sel-count{color:#c62828;font-weight:600;}
.batchbar.hidden{display:none;}
.cards{padding:14px 22px;display:grid;grid-template-columns:repeat(auto-fill,minmax(440px,1fr));gap:12px;}
.chap-head{grid-column:1/-1;display:flex;align-items:center;gap:10px;margin:6px 0 2px;}
.chap-head h2{font-size:15px;color:#1a3c6e;margin:0;}
.chap-head .cnt{font-size:12px;color:#888;}
.chap-head .select-all{margin-left:auto;font-size:11px;color:#2c5aa0;cursor:pointer;text-decoration:underline;}
.card{background:#fff;border-radius:6px;border:1px solid #e0e0e0;border-left:4px solid #ccc;padding:10px 12px;display:flex;flex-direction:column;gap:6px;}
.card.s-pass{border-left-color:#2e7d32;} .card.s-fail{border-left-color:#c62828;}
.card.s-manual{border-left-color:#ef6c00;} .card.s-na{border-left-color:#757575;}
.card.sel{box-shadow:0 0 0 2px #2c5aa0;}
.card-top{display:flex;align-items:flex-start;gap:8px;}
.card-top .chk{margin-top:3px;cursor:pointer;}
.card-top .id{font-weight:700;color:#1a3c6e;font-size:13px;white-space:nowrap;}
.card-top .title{font-size:13px;line-height:1.5;}
.meta{font-size:11px;color:#999;display:flex;gap:12px;flex-wrap:wrap;}
.meta .g{color:#6a8;}
.meta .p{color:#aaa;max-width:100%;}
.res{display:flex;gap:4px;flex-wrap:wrap;}
.res button{border:1px solid #ddd;background:#fff;padding:4px 9px;border-radius:3px;cursor:pointer;font-size:12px;}
.res button.r-pass.sel{background:#2e7d32;color:#fff;border-color:#2e7d32;}
.res button.r-fail.sel{background:#c62828;color:#fff;border-color:#c62828;}
.res button.r-manual.sel{background:#ef6c00;color:#fff;border-color:#ef6c00;}
.res button.r-na.sel{background:#757575;color:#fff;border-color:#757575;}
.field label{font-size:11px;color:#888;display:block;margin-bottom:2px;}
.empty{padding:40px;text-align:center;color:#999;}
footer{padding:12px 22px;color:#888;font-size:11px;text-align:center;}
</style>
</head>
<body>
<header>
  <h1>配置核查 · 人工核查系统</h1>
  <div class="actions">
    <button onclick="saveProgress()">保存进度</button>
    <button onclick="exportExcel()">导出 Excel 报告</button>
    <button class="danger" onclick="clearAll()">清空</button>
  </div>
</header>
<div class="bar">
  <label>被核查系统：<input id="m-os" placeholder="如 银河麒麟V10"></label>
  <label>主机名：<input id="m-host" placeholder="如 lalala-pc"></label>
  <label>核查人：<input id="m-checker" placeholder="姓名"></label>
  <label>核查日期：<input id="m-date" type="date"></label>
  <label>单位：<input id="m-unit" placeholder="可选"></label>
</div>
<div class="stats" id="stats"></div>
<div class="banner" id="banner"></div>
<div class="filters" id="filters">
  <input type="text" id="search" placeholder="搜索核查项关键词…" oninput="render()">
</div>
<div class="batchbar hidden" id="batchbar">
  <span class="lbl">批量操作：</span>
  <span class="sel-count" id="selcount">已选 0 项</span>
  <button onclick="applyBatch('pass')">设为合规</button>
  <button onclick="applyBatch('fail')">设为不合规</button>
  <button onclick="applyBatch('manual')">设为需人工</button>
  <button onclick="applyBatch('na')">设为不适用</button>
  <button onclick="clearResults()">清除结果</button>
  <button onclick="clearSel()">取消选择</button>
</div>
<div class="cards" id="cards"></div>
<footer>人工核查系统 · 数据源自《配置核查表 v2.0.0》共 ITEMS_COUNT 项 · 进度自动保存于本机浏览器</footer>
<script>
const ITEMS = __ITEMS__;
const STAT_CN = {pass:"合规", fail:"不合规", manual:"需人工核查", na:"不适用"};
const STAT_ORDER = ["fail","manual","na","pass"];
const STATE_KEY = "peizhihecha_manual_v1";

let state = loadProgress();
let filterChap = "0";
let selected = new Set();

function loadProgress(){
  try{ const s = JSON.parse(localStorage.getItem(STATE_KEY)||"{}"); if(s.items) return s; }catch(e){}
  return { items:{}, meta:{} };
}
function persist(){
  state.meta = { os:val("m-os"), host:val("m-host"), checker:val("m-checker"), date:val("m-date"), unit:val("m-unit") };
  localStorage.setItem(STATE_KEY, JSON.stringify(state));
}
function val(id){ return document.getElementById(id).value||""; }
function setVal(id,v){ if(v) document.getElementById(id).value=v; }

setVal("m-os", state.meta.os); setVal("m-host", state.meta.host);
setVal("m-checker", state.meta.checker); setVal("m-date", state.meta.date);
setVal("m-unit", state.meta.unit);
["m-os","m-host","m-checker","m-date","m-unit"].forEach(id=>{
  document.getElementById(id).addEventListener("input", persist);
});

function itemState(id){ return state.items[id] || (state.items[id]={result:"",detail:"",rec:""}); }
function setResult(id, r){
  const it = itemState(id); it.result = (it.result===r) ? "" : r; persist(); render();
}
function setDetail(id, v){ itemState(id).detail=v; persist(); updateStats(); }
function setRec(id, v){ itemState(id).rec=v; persist(); }

function buildFilters(){
  const counts = {}, chapNames = {};
  ITEMS.forEach(i=>{ counts[i.cno]=(counts[i.cno]||0)+1; if(!chapNames[i.cno]) chapNames[i.cno]=i.chapter; });
  const f = document.getElementById("filters");
  f.querySelectorAll(".tab").forEach(e=>e.remove());
  const all = document.createElement("span");
  all.className="tab"+(filterChap==="0"?" active":"");
  all.innerHTML="全部 <span class='n'>"+ITEMS.length+"</span>";
  all.onclick=()=>{filterChap="0";buildFilters();render();};
  f.insertBefore(all, f.firstChild);
  Object.keys(chapNames).sort((a,b)=>+a-+b).forEach(cno=>{
    const t=document.createElement("span");
    t.className="tab"+(filterChap===cno?" active":"");
    t.innerHTML=chapNames[cno]+" <span class='n'>"+counts[cno]+"</span>";
    t.onclick=()=>{filterChap=cno;buildFilters();render();};
    f.insertBefore(t, document.getElementById("search"));
  });
}

function filteredItems(){
  const q = document.getElementById("search").value.trim().toLowerCase();
  return ITEMS.filter(i=>{
    if(filterChap!=="0" && String(i.cno)!==filterChap) return false;
    if(q && !(i.title.toLowerCase().includes(q)||i.id.toLowerCase().includes(q))) return false;
    return true;
  });
}

function render(){
  const rows = filteredItems();
  const wrap = document.getElementById("cards");
  if(!rows.length){ wrap.innerHTML='<div class="empty">无匹配项</div>'; updateStats(); return; }
  const groups = {};
  rows.forEach(i=>{ (groups[i.cno] = groups[i.cno]||[]).push(i); });
  let html = "";
  Object.keys(groups).sort((a,b)=>+a-+b).forEach(cno=>{
    const arr = groups[cno];
    const allSel = arr.every(i=>selected.has(i.id));
    html += `<div class="chap-head"><h2>第${cno}章 ${arr[0].chapter}</h2><span class="cnt">${arr.length} 项</span><span class="select-all" onclick="toggleChapSel('${cno}', ${allSel?'true':'false'})">${allSel?'取消全选':'全选本章'}</span></div>`;
    html += arr.map(i=>{
      const st = itemState(i.id);
      const mkBtn = (r)=>`<button class="r-${r} ${st.result===r?'sel':''}" onclick="setResult('${i.id}','${r}')">${STAT_CN[r]}</button>`;
      const chk = selected.has(i.id) ? "checked" : "";
      return `<div class="card s-${st.result||'none'} ${selected.has(i.id)?'sel':''}">
        <div class="card-top">
          <input type="checkbox" class="chk" ${chk} onchange="toggleSel('${i.id}',this.checked)">
          <span class="id">${i.id}</span>
          <span class="title">${esc(i.title)}</span>
        </div>
        <div class="meta">
          <span class="g">📖 ${esc(i.guide)}</span>
          ${i.plat?`<span class="p">适用：${esc(i.plat)}</span>`:''}
        </div>
        <div class="res">${["pass","fail","manual","na"].map(mkBtn).join("")}</div>
        <div class="field"><label>核查发现 / 证据</label><textarea rows="2" oninput="setDetail('${i.id}',this.value)" placeholder="填写现场核查发现">${esc(st.detail)}</textarea></div>
        <div class="field"><label>整改建议</label><textarea rows="2" oninput="setRec('${i.id}',this.value)" placeholder="整改建议">${esc(st.rec)}</textarea></div>
      </div>`;
    }).join("");
  });
  wrap.innerHTML = html;
  updateStats();
}

function toggleSel(id, on){ if(on) selected.add(id); else selected.delete(id); render(); }
function toggleChapSel(cno, unselect){
  filteredItems().filter(i=>String(i.cno)===cno).forEach(i=>{
    if(unselect) selected.delete(i.id); else selected.add(i.id);
  });
  render();
}
function clearSel(){ selected.clear(); render(); }
function applyBatch(r){ if(!selected.size) return; selected.forEach(id=>{ itemState(id).result=r; }); persist(); render(); }
function clearResults(){ if(!selected.size) return; if(!confirm("清除选中项的结果？")) return; selected.forEach(id=>{ itemState(id).result=""; }); persist(); render(); }

function updateStats(){
  let pass=0,fail=0,manual=0,na=0,empty=0;
  ITEMS.forEach(i=>{
    const r=itemState(i.id).result;
    if(r==="pass")pass++; else if(r==="fail")fail++; else if(r==="manual")manual++; else if(r==="na")na++; else empty++;
  });
  const filled = ITEMS.length-empty;
  document.getElementById("stats").innerHTML =
    `<span class="chip s-pass">合规 ${pass}</span><span class="chip s-fail">不合规 ${fail}</span>`+
    `<span class="chip s-manual">需人工 ${manual}</span><span class="chip s-na">不适用 ${na}</span>`+
    `<span class="chip s-empty">未填 ${empty}</span>`+
    `<span class="progress">已核查 <b>${filled}/${ITEMS.length}</b></span>`;
  const banner = document.getElementById("banner");
  if(empty>0){ banner.className="banner show"; banner.innerHTML=`⚠ 还有 ${empty} 项未填写结果，导出的报告将不含这些项。`; }
  else { banner.className="banner"; }
  const n=selected.size;
  document.getElementById("selcount").textContent="已选 "+n+" 项";
  document.getElementById("batchbar").className="batchbar"+(n?"":" hidden");
}

function esc(s){ return String(s==null?"":s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;"); }

function saveProgress(){ persist();
  const blob = new Blob([JSON.stringify(state,null,2)],{type:"application/json"});
  const a=document.createElement("a"); a.href=URL.createObjectURL(blob);
  a.download=`核查进度_${stamp()}.json`; a.click(); URL.revokeObjectURL(a.href);
  alert("进度已保存（本机+下载了一份json备份）");
}
function clearAll(){
  if(!confirm("确定清空所有核查结果？此操作不可撤销。")) return;
  state={items:{},meta:{}}; localStorage.removeItem(STATE_KEY); selected.clear();
  ["m-os","m-host","m-checker","m-date","m-unit"].forEach(id=>document.getElementById(id).value="");
  render(); buildFilters();
}
function stamp(){
  const d=new Date(); const p=n=>String(n).padStart(2,"0");
  return `${d.getFullYear()}${p(d.getMonth()+1)}${p(d.getDate())}_${p(d.getHours())}${p(d.getMinutes())}${p(d.getSeconds())}`;
}

function exportExcel(){
  persist();
  const filled = ITEMS.filter(i=>itemState(i.id).result);
  if(!filled.length){ alert("尚未填写任何结果，无法导出。"); return; }
  const byStatus = {}; STAT_ORDER.forEach(s=>byStatus[s]=[]);
  filled.forEach(i=>byStatus[itemState(i.id).result].push(i));
  const sec = [["fail","一、未通过"],["manual","二、需人工核查"],["na","三、不适用"],["pass","四、通过"]];
  const escH = s=>String(s==null?"":s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
  let rows="";
  sec.forEach(([k,title])=>{
    const arr=byStatus[k]; if(!arr.length) return;
    rows += `<p><b>${title}（${arr.length} 项）</b></p>`;
    rows += `<table border=1><tr style="background:#1a3c6e;color:#fff;"><th>章节</th><th>编号</th><th>类别</th><th>核查项</th><th>结果</th><th>详情</th><th>建议</th><th>参考指导书</th></tr>`;
    arr.forEach(i=>{
      const st=itemState(i.id);
      rows += `<tr><td>${escH(i.chapter)}</td><td>${escH(i.id)}</td><td>${escH(i.category)}</td><td>${escH(i.title)}</td><td>${escH(STAT_CN[st.result])}</td><td>${escH(st.detail)}</td><td>${escH(st.rec)}</td><td>${escH(i.guide)}</td></tr>`;
    });
    rows += `</table>`;
  });
  const m=state.meta;
  const cnt = {pass:byStatus.pass.length,fail:byStatus.fail.length,manual:byStatus.manual.length,na:byStatus.na.length};
  const html = `<html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel" xmlns="http://www.w3.org/TR/REC-html40">
<head><meta charset="UTF-8"><style>table{border-collapse:collapse;mso-number-format:"\@";}th,td{border:1px solid #999;padding:4px 6px;font-family:"Microsoft YaHei";font-size:12px;}th{background:#1a3c6e;color:#fff;}p{margin:8px 0;}</style></head>
<body>
<p style="font-size:14px;font-weight:bold;">配置核查报告（人工核查版）</p>
<p>被核查系统：${escH(m.os)}　主机名：${escH(m.host)}　核查人：${escH(m.checker)}　核查日期：${escH(m.date)}　单位：${escH(m.unit)}</p>
<p>合规：${cnt.pass}　不合规：${cnt.fail}　需人工核查：${cnt.manual}　不适用：${cnt.na}　（共 ${filled.length} 项）</p>
${rows}
</body></html>`;
  const blob = new Blob(["\ufeff"+html],{type:"application/vnd.ms-excel;charset=utf-8"});
  const a=document.createElement("a"); a.href=URL.createObjectURL(blob);
  a.download=`配置核查报告_人工核查_${stamp()}.xls`; a.click(); URL.revokeObjectURL(a.href);
}

buildFilters();
render();
</script>
</body>
</html>
'''


def main():
    items = build_items()
    html = HTML.replace("__ITEMS__", json.dumps(items, ensure_ascii=False))
    html = html.replace("ITEMS_COUNT", str(len(items)))
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"saved: {OUT}")
    print(f"items: {len(items)}")
    from collections import Counter
    c = Counter(it["chapter"] for it in items)
    for ch, n in c.items():
        print(f"  {ch}: {n}")


if __name__ == "__main__":
    main()
