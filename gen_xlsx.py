#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配置核查报告 HTML -> 真·Excel(.xlsx) 转换器 / 汇总合并器

脚本本身已经会输出 .html 和 .xls（后者其实是 HTML 表格改扩展名，Excel/WPS 能打开但会提示格式不符）。
本工具把 HTML 报告转换成真正的 .xlsx 二进制文件，并把一次 run_all 产生的多个组件报告合并成一个汇总 Excel。

用法：
    python gen_xlsx.py                 # 递归转换项目下所有 output/ 目录中的 .html
    python gen_xlsx.py --merge         # 把各组件最新报告合并成一个 配置核查汇总报告_*.xlsx
    python gen_xlsx.py 某个.html       # 转换单个文件
    python gen_xlsx.py 某个目录        # 转换该目录及其 output/ 子目录下的 .html

兼容两类报告格式：
    bash/PowerShell 脚本（8 列：章节/编号/类别/核查项/结果/详情/建议/参考指导书）
    VBScript 脚本（Windows XP/7，列：编号/类别/检查项/详情/修复建议/章节，GBK 编码）

依赖：openpyxl（pip install openpyxl）
"""
import os
import re
import sys
import glob
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("缺少依赖 openpyxl，请先执行：pip install openpyxl")
    sys.exit(1)


HEADERS = ["章节", "编号", "类别", "核查项", "结果", "详情", "建议", "参考指导书"]

STATUS_LABEL = {"fail": "未通过", "manual": "需人工核查", "pass": "合规", "na": "不适用"}
STATUS_ORDER = ["fail", "manual", "pass", "na"]

RESULT_FILL = {"fail": "FFC7CE", "manual": "FFEB9C", "pass": "C6EFCE", "na": "D9D9D9"}
RESULT_FONT = {"fail": "9C0006", "manual": "9C6500", "pass": "006100", "na": "404040"}


def status_key(text):
    """把结果文本映射为 fail/manual/pass/na/None。"""
    if text is None:
        return None
    t = str(text)
    if any(k in t for k in ("不合规", "未通过", "失败")):
        return "fail"
    if any(k in t for k in ("人工", "需人工")):
        return "manual"
    if any(k in t for k in ("合规", "通过")):
        return "pass"
    if "不适用" in t:
        return "na"
    return None


def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = (s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
          .replace("&quot;", '"').replace("&#39;", "'").replace("&nbsp;", " ")
          .replace("\u00a0", " "))
    return s.strip()


def read_html(path):
    """读取 HTML，自动识别 UTF-8 / GBK 编码（win_xp7 的 VBScript 输出为 GBK）。"""
    with open(path, "rb") as f:
        raw = f.read()
    for enc in ("utf-8-sig", "utf-8", "gbk", "cp936"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def parse_meta(html):
    m = re.search(r'<div class="meta">(.*?)<div class="summary">', html, re.S)
    if not m:
        return []
    return [strip_tags(x) for x in re.findall(r"<div>(.*?)</div>", m.group(1), re.S)]


def parse_summary(html):
    s = {}
    # 格式1（bash/PowerShell）：<div class="card pass">合规<br>0</div>
    for m in re.findall(r'class="card (pass|fail|manual|na)">\s*([^<]*)<br>(\d+)', html):
        s[m[0]] = (m[1], int(m[2]))
    # 格式2（VBScript）：<p class="pass"><strong>通过：</strong>16 项</p>
    if not s:
        for m in re.findall(r'class="(pass|fail|manual|na)"[^>]*>\s*<strong>([^<]*?)[：:]?</strong>\s*(\d+)', html):
            s[m[0]] = (m[1], int(m[2]))
    return s


def infer_status(title):
    return status_key(title)


def extract_tables(content):
    tables = []
    for tm in re.findall(r"<table>(.*?)</table>", content, re.S):
        rows = []
        for tr in re.findall(r"<tr>(.*?)</tr>", tm, re.S):
            cells = [strip_tags(td) for td in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S)]
            if not cells:
                cells = [strip_tags(th) for th in re.findall(r"<th[^>]*>(.*?)</th>", tr, re.S)]
            if cells:
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables


def parse_tables(html):
    """返回 [(section_title, status, [ [row,...], ... ]), ...]，status 来自 h2 class 或标题。"""
    sections = []
    parts = re.split(r"<h2([^>]*)>(.*?)</h2>", html, flags=re.S)
    for i in range(1, len(parts) - 1, 3):
        attrs = parts[i]
        title = strip_tags(parts[i + 1])
        content = parts[i + 2]
        status = None
        cm = re.search(r'class=["\']([^"\']+)', attrs)
        if cm:
            status = status_key(cm.group(1))
        if not status:
            status = infer_status(title)
        sections.append((title, status, extract_tables(content)))
    return sections


def report_data(html_path):
    html = read_html(html_path)
    title = strip_tags(re.search(r"<h1>(.*?)</h1>", html, re.S).group(1)) if re.search(r"<h1>(.*?)</h1>", html, re.S) else "配置核查报告"
    return title, parse_meta(html), parse_summary(html), parse_tables(html)


def write_detail_sheet(wd, sections):
    """把明细写入工作表，含动态表头 + 分类列 + 结果着色。"""
    header_fill = PatternFill("solid", fgColor="1A3C6E")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header = None
    for _, _, tables in sections:
        for rows in tables:
            if rows and len(rows[0]) > 1:
                header = rows[0]
                break
        if header:
            break
    if not header:
        header = HEADERS
    all_headers = ["分类"] + list(header)
    result_col = None
    for idx, h in enumerate(header):
        if h == "结果":
            result_col = idx + 2
            break

    for c, h in enumerate(all_headers, 1):
        cell = wd.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    row_idx = 2
    for sec_title, status, tables in sections:
        cat_label = STATUS_LABEL.get(status, re.sub(r"^[一二三四五六七八九十]+、", "", sec_title))
        for rows in tables:
            for r0 in rows:
                if not r0 or r0 == header:
                    continue
                data = list(r0[:len(header)])
                while len(data) < len(header):
                    data.append("")
                cat_cell = wd.cell(row=row_idx, column=1, value=cat_label)
                cat_cell.alignment = Alignment(vertical="top")
                cat_cell.border = border
                if status:
                    cat_cell.fill = PatternFill("solid", fgColor=RESULT_FILL[status])
                    cat_cell.font = Font(bold=True, color=RESULT_FONT[status])
                for c, v in enumerate(data, 2):
                    cell = wd.cell(row=row_idx, column=c, value=v)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border
                if result_col:
                    res_cell = wd.cell(row=row_idx, column=result_col)
                    k = status_key(res_cell.value)
                    if k:
                        res_cell.fill = PatternFill("solid", fgColor=RESULT_FILL[k])
                        res_cell.font = Font(bold=True, color=RESULT_FONT[k])
                row_idx += 1

    ncol = len(all_headers)
    default_widths = [10, 8, 8, 12, 26, 10, 46, 40, 40]
    for i in range(1, ncol + 1):
        w = default_widths[i - 1] if i - 1 < len(default_widths) else 20
        wd.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wd.freeze_panes = "A2"
    wd.auto_filter.ref = "A1:%s%d" % (openpyxl.utils.get_column_letter(ncol), row_idx - 1)


def html_to_xlsx(html_path, xlsx_path):
    title, meta, summary, sections = report_data(html_path)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "汇总"
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="1A3C6E")
    r = 3
    for line in meta:
        ws.cell(row=r, column=1, value=line).font = Font(size=11)
        r += 1
    r += 1
    labels = {"pass": "合规", "fail": "不合规", "manual": "需人工核查", "na": "不适用"}
    for key, label in labels.items():
        val = summary.get(key, (label, 0))
        cell = ws.cell(row=r, column=1, value="%s：%d 项" % (val[0], val[1]))
        cell.font = Font(bold=True, size=12)
        r += 1
    ws.column_dimensions["A"].width = 60

    wd = wb.create_sheet("核查明细")
    write_detail_sheet(wd, sections)

    wb.save(xlsx_path)
    return xlsx_path


# ---------- 合并汇总 ----------

def component_from_filename(path):
    """从文件名提取组件名（如 配置核查报告_MySQL_20260815_... -> MySQL）；OS 报告返回 None。"""
    base = os.path.basename(path)
    name = base.replace("配置核查报告_", "").replace(".html", "")
    parts = name.split("_")
    if len(parts) >= 2 and not parts[0].isdigit():
        return parts[0]
    return None


def os_name_from_title(title):
    if "麒麟" in title:
        return "麒麟OS"
    if "Windows" in title or "XP" in title or "XP/7" in title:
        return "WinOS"
    return "操作系统"


def sheet_key(path, title):
    comp = component_from_filename(path)
    if comp:
        return re.sub(r'[\\/*?:\[\]]', '', comp)[:31]
    return os_name_from_title(title)[:31]


def merge_reports(targets, out_path):
    """把各组件最新报告合并成一个工作簿：总览 + 每组件一个 sheet。"""
    # 按组件分组，取每个组件最新的一份
    groups = {}
    for t in targets:
        try:
            title = report_data(t)[0]
        except Exception:
            continue
        key = sheet_key(t, title)
        mtime = os.path.getmtime(t)
        if key not in groups or mtime > groups[key][0]:
            groups[key] = (mtime, t)

    wb = openpyxl.Workbook()
    ov = wb.active
    ov.title = "总览"
    ov.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1A3C6E")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ov_headers = ["组件", "合规", "不合规", "需人工核查", "不适用", "总项数"]
    for c, h in enumerate(ov_headers, 1):
        cell = ov.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = 2
    for key in sorted(groups):
        mtime, t = groups[key]
        title, meta, summary, sections = report_data(t)
        counts = {k: summary.get(k, (None, 0))[1] for k in ("pass", "fail", "manual", "na")}
        total = sum(counts.values())
        vals = [key, counts.get("pass", 0), counts.get("fail", 0), counts.get("manual", 0), counts.get("na", 0), total]
        for c, v in enumerate(vals, 1):
            cell = ov.cell(row=row, column=c, value=v)
            cell.border = border
            if c >= 2:
                cell.alignment = Alignment(horizontal="center")
        row += 1

        ws = wb.create_sheet(key)
        write_detail_sheet(ws, sections)

    for i, w in enumerate([16, 8, 8, 10, 8, 8], 1):
        ov.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ov.freeze_panes = "A2"

    wb.save(out_path)
    return out_path


def collect_targets(args):
    targets = []
    for a in args:
        if os.path.isfile(a):
            if a.lower().endswith(".html"):
                targets.append(a)
        elif os.path.isdir(a):
            targets.extend(sorted(glob.glob(os.path.join(a, "*.html"))))
            targets.extend(sorted(glob.glob(os.path.join(a, "output", "*.html"))))
    if not args:
        base = os.path.dirname(os.path.abspath(__file__))
        for d in ("kylin/output", "win/output"):
            p = os.path.join(base, d)
            if os.path.isdir(p):
                targets.extend(sorted(glob.glob(os.path.join(p, "*.html"))))
    return targets


def main():
    args = sys.argv[1:]
    do_merge = "--merge" in args
    args = [a for a in args if a != "--merge"]

    targets = collect_targets(args)
    if not targets:
        print("未找到 .html 报告文件。用法：python gen_xlsx.py [html文件或目录] [--merge]")
        sys.exit(0)

    if do_merge:
        stamp = time.strftime("%Y%m%d_%H%M%S")
        out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "配置核查汇总报告_%s.xlsx" % stamp)
        try:
            merge_reports(targets, out)
            print("已生成汇总报告：%s" % out)
        except Exception as e:
            print("合并失败：%s" % e)
        return

    made = 0
    for h in targets:
        x = os.path.splitext(h)[0] + ".xlsx"
        try:
            html_to_xlsx(h, x)
            made += 1
            print("已生成：%s" % os.path.basename(x))
        except Exception as e:
            print("转换失败 %s：%s" % (h, e))
    print("完成，共转换 %d 个报告。" % made)


if __name__ == "__main__":
    main()
